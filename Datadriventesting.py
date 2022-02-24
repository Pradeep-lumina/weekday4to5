import time

import openpyxl

from selenium import webdriver
from selenium.webdriver.common.keys import Keys

driver = webdriver.Chrome(executable_path="D:\chromedriver_win32 (11)\chromedriver.exe")

driver.get("https://demo.guru99.com/test/newtours/index.php")

driver.maximize_window()

path = "D:\Fita class\Selenium Python\\Values.xlsx"

workbook = openpyxl.load_workbook(path)

sheet = workbook.get_sheet_by_name("Sheet1")

rows = sheet.max_row
columns = sheet.max_column

print(rows)
print(columns)

for r in range(2, rows+1):
    username = sheet.cell(r, 1).value
    password = sheet.cell(r, 2).value

    driver.find_element_by_name("userName").send_keys(username)

    driver.find_element_by_name("password").send_keys(password)

    driver.find_element_by_name("submit").click()

    if driver.title == "Login: Mercury Tours":
        print("Test Passed")
        sheet.cell(r, 3, "Test Passed")
        driver.find_element_by_xpath("//a[text()='SIGN-OFF']").click()
        time.sleep(10)
    else:
        print("Test Failed")
        sheet.cell(r, 3, "Test Failed")

    workbook.save(path)